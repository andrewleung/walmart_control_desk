from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import csv

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import ExceptionRecord, PackageUpload, ReconciliationRow, SKU


def _clean(value):
    return str(value).strip() if value is not None else ""


def _number(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _yes(value):
    return _clean(value).upper() in {"YES", "Y", "TRUE", "1", "ON TIME", "IN FULL", "PASS"}


def _iter_rows(path):
    path = Path(path)
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            yield from csv.DictReader(handle)
        return
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [_clean(v) for v in next(iterator)]
    try:
        for row in iterator:
            yield {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
    finally:
        workbook.close()


def _sku(vendor_stock_id, item_name="", walmart_item_number=""):
    sku, _ = SKU.objects.get_or_create(
        vendor_stock_id=vendor_stock_id,
        defaults={
            "item_name": item_name or vendor_stock_id,
            "walmart_item_number": _clean(walmart_item_number),
        },
    )
    changed = False
    if item_name and not sku.item_name:
        sku.item_name, changed = item_name, True
    if walmart_item_number and not sku.walmart_item_number:
        sku.walmart_item_number, changed = _clean(walmart_item_number), True
    if changed:
        sku.save()
    return sku


def _row(cycle, sku):
    row, _ = ReconciliationRow.objects.get_or_create(cycle=cycle, sku=sku)
    return row


@transaction.atomic
def process_upload(upload):
    upload.status = "PROCESSING"
    upload.validation_message = ""
    upload.save(update_fields=["status", "validation_message"])
    try:
        handlers = {
            1: _package_1,
            2: _package_2,
            3: _package_3,
            4: _package_4,
            5: _package_5,
            6: _package_6,
            7: _package_7,
            8: _package_8,
            9: _package_9,
        }
        handler = handlers.get(upload.package_number)
        if handler:
            count = handler(upload)
            upload.status = "POPULATED"
            upload.validation_message = f"Validated and loaded {count:,} source rows."
            upload.row_count = count
        else:
            upload.status = "PARTIAL"
            upload.validation_message = (
                "File retained. Automated normalization for this package is not included in the MVP; "
                "its contents must be reviewed before reconciliation."
            )
    except Exception as exc:
        upload.status = "ERROR"
        upload.validation_message = str(exc)
    upload.save(update_fields=["status", "validation_message", "row_count"])
    refresh_cycle(upload.cycle)
    return upload


def _package_1(upload):
    grouped = defaultdict(lambda: defaultdict(float))
    instock = defaultdict(list)
    identity = {}
    count = 0
    required = {
        "vendor_stock_id", "walmart_item_number", "item_name",
        "pos_quantity_this_year", "store_on_hand_quantity_this_year",
    }
    observed = set()
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        stock_id = _clean(source.get("vendor_stock_id"))
        if not stock_id:
            continue
        identity[stock_id] = (
            _clean(source.get("item_name")),
            _clean(source.get("walmart_item_number")),
        )
        for source_name, destination in [
            ("pos_quantity_this_year", "store_pos_units"),
            ("pos_sales_this_year", "store_pos_sales"),
            ("store_on_hand_quantity_this_year", "store_on_hand"),
            ("store_on_order_quantity_this_year", "store_on_order"),
        ]:
            grouped[stock_id][destination] += _number(source.get(source_name))
        value = source.get("repl_instock_percentage_this_year")
        if isinstance(value, (int, float)):
            instock[stock_id].append(float(value))
        if source.get("actual_order_quantity_last_4_weeks") not in (None, ""):
            grouped[stock_id]["actual_orders_l4"] += _number(
                source.get("actual_order_quantity_last_4_weeks")
            )
    missing = sorted(required - observed)
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))
    for stock_id, values in grouped.items():
        sku = _sku(stock_id, *identity[stock_id])
        row = _row(upload.cycle, sku)
        for field, value in values.items():
            setattr(row, field, Decimal(str(round(value, 4))))
        if instock[stock_id]:
            row.replen_instock = Decimal(str(sum(instock[stock_id]) / len(instock[stock_id])))
        row.save()
    return count


def _package_2(upload):
    grouped = defaultdict(lambda: defaultdict(float))
    identity = {}
    count = 0
    observed = set()
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        stock_id = _clean(source.get("vendor_stock_id"))
        if not stock_id:
            continue
        identity[stock_id] = (
            _clean(source.get("item_name")),
            _clean(source.get("walmart_item_number")),
        )
        grouped[stock_id]["ecomm_units"] += _number(source.get("shipped_based_quantity_this_year"))
        grouped[stock_id]["ecomm_sales"] += _number(source.get("shipped_based_net_sales_amount_this_year"))
    needed = {"vendor_stock_id", "shipped_based_quantity_this_year", "shipped_based_net_sales_amount_this_year"}
    if missing := sorted(needed - observed):
        raise ValueError("Missing required columns: " + ", ".join(missing))
    for stock_id, values in grouped.items():
        row = _row(upload.cycle, _sku(stock_id, *identity[stock_id]))
        row.ecomm_units = Decimal(str(round(values["ecomm_units"], 4)))
        row.ecomm_sales = Decimal(str(round(values["ecomm_sales"], 4)))
        row.save()
    return count


def _recalculate_usable_supply(row):
    parts = [
        row.ecomm_available_inventory,
        row.factory_available_inventory,
        row.rjw_available_inventory,
    ]
    if any(value is not None for value in parts):
        row.usable_supply = sum((value or Decimal("0")) for value in parts)


def _package_3(upload):
    grouped = defaultdict(float)
    on_hand = defaultdict(float)
    facilities = defaultdict(set)
    count = 0
    observed = set()
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        stock_id = _clean(source.get("vendor_stock_id"))
        if stock_id:
            grouped[stock_id] += _number(source.get("fc_available_quantity"))
            on_hand[stock_id] += _number(source.get("fc_on_hand_quantity"))
            facilities[stock_id].add(_clean(source.get("fc_number")))
    needed = {"vendor_stock_id", "fc_number", "fc_on_hand_quantity", "fc_available_quantity"}
    if missing := sorted(needed - observed):
        raise ValueError("Missing required columns: " + ", ".join(missing))
    for stock_id, total in grouped.items():
        row = _row(upload.cycle, _sku(stock_id))
        row.ecomm_available_inventory = Decimal(str(round(total, 4)))
        row.ecomm_on_hand_inventory = Decimal(str(round(on_hand[stock_id], 4)))
        row.ecomm_fc_count = len({value for value in facilities[stock_id] if value})
        if row.ecomm_units:
            row.ecomm_weeks_of_supply = row.ecomm_on_hand_inventory / row.ecomm_units
        _recalculate_usable_supply(row)
        row.save()
    return count


def _package_4(upload):
    grouped = defaultdict(lambda: defaultdict(float))
    count = 0
    observed = set()
    wm_to_sku = {s.walmart_item_number: s for s in SKU.objects.exclude(walmart_item_number="")}
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        wm = _clean(source.get("wm_item_nbr"))
        sku = wm_to_sku.get(wm)
        if not sku:
            continue
        week = _clean(source.get("walmart_calendar_week"))
        grouped[sku.pk][week] += _number(source.get("final_forecast_each_quantity"))
    needed = {"wm_item_nbr", "walmart_calendar_week", "final_forecast_each_quantity"}
    if missing := sorted(needed - observed):
        raise ValueError("Missing required columns: " + ", ".join(missing))
    for sku_id, weeks in grouped.items():
        ordered = sorted(k for k in weeks if k)
        row = _row(upload.cycle, SKU.objects.get(pk=sku_id))
        row.forecast_demand_4w = Decimal(str(round(sum(weeks[w] for w in ordered[:4]), 4)))
        row.forecast_demand_13w = Decimal(str(round(sum(weeks.values()), 4)))
        row.save()
    return count


def _package_5(upload):
    grouped = defaultdict(float)
    arrivals = defaultdict(list)
    supply_plan = defaultdict(lambda: defaultdict(float))
    identity = {}
    count = 0
    observed = set()
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        stock_id = _clean(source.get("vendor_stock_id"))
        if not stock_id:
            continue
        identity[stock_id] = (
            _clean(source.get("item_name")),
            _clean(source.get("wm_item_nbr")),
        )
        grouped[stock_id] += _number(source.get("order_each_quantity"))
        week = _clean(source.get("walmart_calendar_week"))
        if week and source.get("supply_plan_each_quantity") not in (None, ""):
            supply_plan[stock_id][week] += _number(source.get("supply_plan_each_quantity"))
        if arrival := _date(source.get("sched_arvl_dt")):
            arrivals[stock_id].append(arrival)
    needed = {"vendor_stock_id", "order_each_quantity", "sched_arvl_dt"}
    if missing := sorted(needed - observed):
        raise ValueError("Missing required columns: " + ", ".join(missing))
    for stock_id, total in grouped.items():
        row = _row(upload.cycle, _sku(stock_id, *identity[stock_id]))
        row.order_forecast_total = Decimal(str(round(total, 4)))
        if supply_plan[stock_id]:
            first_week = sorted(supply_plan[stock_id])[0]
            row.next_week_supply_plan = Decimal(str(round(supply_plan[stock_id][first_week], 4)))
        if arrivals[stock_id]:
            row.first_forecast_arrival = min(arrivals[stock_id])
            row.last_forecast_arrival = max(arrivals[stock_id])
        row.save()
    return count


def _package_6(upload):
    commitments = defaultdict(float)
    mabds = defaultdict(list)
    otif = defaultdict(lambda: {"rows": 0, "on_time": 0, "in_full": 0, "exceptions": 0})
    count = 0
    observed = set()
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        stock_id = _clean(source.get("vendor_stock_id"))
        if not stock_id:
            continue
        commitments[stock_id] += _number(source.get("unreceived_quantity"))
        if source.get("otif_on_time_flag") not in (None, ""):
            otif[stock_id]["rows"] += 1
            otif[stock_id]["on_time"] += int(_yes(source.get("otif_on_time_flag")))
            otif[stock_id]["in_full"] += int(_yes(source.get("otif_in_full_flag")))
            if _clean(source.get("otif_exception_reason")):
                otif[stock_id]["exceptions"] += 1
        if value := _date(source.get("mabd")):
            mabds[stock_id].append(value)
    needed = {
        "po_number", "po_line", "vendor_stock_id", "mabd",
        "ordered_quantity", "received_quantity", "unreceived_quantity",
    }
    if missing := sorted(needed - observed):
        raise ValueError("Missing required columns: " + ", ".join(missing))
    for stock_id, total in commitments.items():
        row = _row(upload.cycle, _sku(stock_id))
        row.current_commitments = Decimal(str(round(total, 4)))
        if mabds[stock_id]:
            row.next_mabd = min(mabds[stock_id])
        if otif[stock_id]["rows"]:
            denominator = Decimal(otif[stock_id]["rows"])
            row.otif_on_time_percent = Decimal(otif[stock_id]["on_time"]) / denominator
            row.otif_in_full_percent = Decimal(otif[stock_id]["in_full"]) / denominator
            row.otif_exception_count = otif[stock_id]["exceptions"]
        row.save()
    return count


def _package_7(upload):
    available = defaultdict(float)
    wip = defaultdict(float)
    completions = defaultdict(list)
    releases = defaultdict(list)
    count = 0
    observed = set()
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        stock_id = _clean(source.get("vendor_stock_id"))
        if stock_id:
            available[stock_id] += _number(source.get("finished_goods_available_quantity"))
            wip[stock_id] += _number(source.get("work_in_process_quantity"))
            if value := _date(source.get("planned_completion_date")):
                completions[stock_id].append(value)
            if value := _date(source.get("release_date")):
                releases[stock_id].append(value)
    needed = {
        "vendor_stock_id", "finished_goods_available_quantity",
        "work_in_process_quantity", "planned_completion_date", "release_date",
    }
    if missing := sorted(needed - observed):
        raise ValueError("Missing required columns: " + ", ".join(missing))
    for stock_id, total in available.items():
        row = _row(upload.cycle, _sku(stock_id))
        row.factory_available_inventory = Decimal(str(round(total, 4)))
        row.work_in_process_quantity = Decimal(str(round(wip[stock_id], 4)))
        row.next_factory_completion = min(completions[stock_id]) if completions[stock_id] else None
        row.next_factory_release = min(releases[stock_id]) if releases[stock_id] else None
        _recalculate_usable_supply(row)
        row.save()
    return count


def _package_8(upload):
    available = defaultdict(float)
    on_time = defaultdict(float)
    late = defaultdict(float)
    inbound_dates = defaultdict(list)
    physical = defaultdict(float)
    allocated = defaultdict(float)
    held = defaultdict(float)
    etds = defaultdict(list)
    etas = defaultdict(list)
    customs = defaultdict(list)
    count = 0
    observed = set()
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        stock_id = _clean(source.get("vendor_stock_id"))
        if not stock_id:
            continue
        available[stock_id] += _number(source.get("rjw_available_quantity"))
        physical[stock_id] += _number(source.get("rjw_physical_quantity"))
        allocated[stock_id] += _number(source.get("rjw_allocated_quantity"))
        held[stock_id] += _number(source.get("rjw_held_quantity"))
        quantity = _number(source.get("inbound_quantity"))
        if _clean(source.get("on_time_for_mabd")).upper() in {"YES", "Y", "TRUE", "1"}:
            on_time[stock_id] += quantity
        else:
            late[stock_id] += quantity
        if value := _date(source.get("expected_rjw_available_date")):
            inbound_dates[stock_id].append(value)
        for source_name, destination in [
            ("etd", etds), ("eta", etas), ("customs_clearance_date", customs)
        ]:
            if value := _date(source.get(source_name)):
                destination[stock_id].append(value)
    needed = {
        "vendor_stock_id", "rjw_physical_quantity", "rjw_available_quantity",
        "rjw_allocated_quantity", "rjw_held_quantity", "inbound_quantity",
        "expected_rjw_available_date", "on_time_for_mabd",
    }
    if missing := sorted(needed - observed):
        raise ValueError("Missing required columns: " + ", ".join(missing))
    for stock_id in available:
        row = _row(upload.cycle, _sku(stock_id))
        row.rjw_available_inventory = Decimal(str(round(available[stock_id], 4)))
        row.rjw_physical_inventory = Decimal(str(round(physical[stock_id], 4)))
        row.rjw_allocated_inventory = Decimal(str(round(allocated[stock_id], 4)))
        row.rjw_held_inventory = Decimal(str(round(held[stock_id], 4)))
        row.confirmed_on_time_inbound = Decimal(str(round(on_time[stock_id], 4)))
        row.late_inbound_quantity = Decimal(str(round(late[stock_id], 4)))
        if inbound_dates[stock_id]:
            row.next_inbound_available = min(inbound_dates[stock_id])
        row.next_etd = min(etds[stock_id]) if etds[stock_id] else None
        row.next_eta = min(etas[stock_id]) if etas[stock_id] else None
        row.next_customs_clearance = min(customs[stock_id]) if customs[stock_id] else None
        _recalculate_usable_supply(row)
        row.save()
    return count


def _package_9(upload):
    count = 0
    observed = set()
    for source in _iter_rows(upload.source_file.path):
        count += 1
        observed.update(source.keys())
        stock_id = _clean(source.get("vendor_stock_id"))
        if not stock_id:
            continue
        row = _row(upload.cycle, _sku(stock_id))
        row.approved_buffer = Decimal(str(round(_number(source.get("approved_buffer_quantity")), 4)))
        row.recommendation = _clean(source.get("required_action"))
        row.modular_set_week = _clean(source.get("modular_set_week"))
        row.system_order_start_week = _clean(source.get("system_order_start_week"))
        row.modular_set_date = _date(source.get("modular_set_date"))
        if source.get("traited_store_count") not in (None, ""):
            row.traited_store_count = int(_number(source.get("traited_store_count")))
        if source.get("prior_traited_store_count") not in (None, ""):
            row.prior_traited_store_count = int(_number(source.get("prior_traited_store_count")))
        if row.traited_store_count is not None and row.prior_traited_store_count is not None:
            row.incremental_store_count = row.traited_store_count - row.prior_traited_store_count
        if source.get("initial_fill_units_per_store") not in (None, ""):
            row.initial_fill_units_per_store = Decimal(
                str(round(_number(source.get("initial_fill_units_per_store")), 4))
            )
        if row.incremental_store_count is not None and row.initial_fill_units_per_store is not None:
            row.illustrative_initial_fill = (
                Decimal(row.incremental_store_count) * row.initial_fill_units_per_store
            )
        row.save()
    needed = {
        "walmart_week", "vendor_stock_id", "context_type", "explanation",
        "required_action", "approved_buffer_quantity", "status",
    }
    if missing := sorted(needed - observed):
        raise ValueError("Missing required columns: " + ", ".join(missing))
    return count


@transaction.atomic
def run_controls(cycle):
    ExceptionRecord.objects.filter(cycle=cycle).delete()
    latest = {}
    for upload in cycle.uploads.all():
        latest.setdefault(upload.package_number, upload)
    critical = {
        3: ("ECOMM_INVENTORY_MISSING", "eCommerce inventory is missing",
            "Total channel inventory and FC availability cannot be assessed.",
            "Download a current SKU/FC inventory report."),
        6: ("COMMITMENTS_MISSING", "Current PO, receipt and OTIF evidence is missing or incomplete",
            "Confirmed commitment quantity and MABD cannot be relied upon.",
            "Upload a current PO/receipt/OTIF report."),
        7: ("FACTORY_SUPPLY_MISSING", "Current factory supply is missing",
            "Finished goods, WIP and availability dates cannot be established.",
            "Upload the current factory production and finished-goods schedule."),
        8: ("RJW_SUPPLY_MISSING", "Current usable supply and inbound timing are missing",
            "Available, allocated and on-time inbound supply cannot be calculated.",
            "Upload current inventory, allocation and inbound timing data."),
    }
    for package, payload in critical.items():
        upload = latest.get(package)
        if not upload or upload.status not in {"POPULATED"}:
            code, title, effect, action = payload
            ExceptionRecord.objects.create(
                cycle=cycle, code=code, severity="CRITICAL", title=title,
                effect=effect, required_action=action,
            )
    if not latest.get(9) or latest[9].status != "POPULATED":
        ExceptionRecord.objects.create(
            cycle=cycle, code="CONTEXT_REQUIRES_UPDATE", severity="HIGH",
            title="Current Walmart context is not validated",
            effect="Forecast or modular exceptions may lack a current explanation.",
            required_action="Add current commentary and exception explanations.",
        )
    for row in cycle.rows.select_related("sku"):
        if row.actual_orders_l4 is not None and row.next_week_supply_plan is not None:
            weekly_average = row.actual_orders_l4 / Decimal("4")
            row.forecast_variance_units = row.next_week_supply_plan - weekly_average
            row.forecast_variance_percent = (
                row.forecast_variance_units / weekly_average if weekly_average else None
            )
            if row.forecast_variance_units < 0:
                ExceptionRecord.objects.create(
                    cycle=cycle, sku=row.sku, code="FORECAST_BELOW_RECENT_ORDERS",
                    severity="HIGH",
                    title=f"{row.sku.vendor_stock_id}: forward supply plan is below recent orders",
                    effect=(
                        f"Next-week supply plan is {abs(row.forecast_variance_units):,.0f} units "
                        "below the last-four-week average order rate."
                    ),
                    required_action=(
                        "Confirm whether recent orders include recovery, modular loading, or other "
                        "one-time quantities before changing the plan."
                    ),
                )
        if row.otif_ready and (
            row.otif_on_time_percent < Decimal("0.90")
            or row.otif_in_full_percent < Decimal("0.90")
        ):
            ExceptionRecord.objects.create(
                cycle=cycle, sku=row.sku, code="OTIF_BELOW_TARGET", severity="HIGH",
                title=f"{row.sku.vendor_stock_id}: OTIF performance is below the review threshold",
                effect=(
                    f"On-time is {row.otif_on_time_percent * 100:.1f}% and "
                    f"in-full is {row.otif_in_full_percent * 100:.1f}%."
                ),
                required_action="Review PO-line exception causes and assign corrective actions.",
            )
        if row.late_inbound_quantity and row.late_inbound_quantity > 0:
            ExceptionRecord.objects.create(
                cycle=cycle, sku=row.sku, code="INBOUND_AFTER_MABD", severity="HIGH",
                title=f"{row.sku.vendor_stock_id}: inbound is not available for the current MABD",
                effect=(
                    f"{row.late_inbound_quantity:,.0f} inbound units are excluded from on-time supply"
                    + (f" and are next expected on {row.next_inbound_available:%d %b %Y}."
                       if row.next_inbound_available else ".")
                ),
                required_action="Escalate timing or cover the commitment from other confirmed supply.",
            )
        if row.missing_critical:
            row.projected_ending_supply = None
            row.projected_gap = None
            row.decision_status = "BLOCKED"
            row.save()
            ExceptionRecord.objects.create(
                cycle=cycle, sku=row.sku, code="SKU_SUPPLY_INCOMPLETE", severity="CRITICAL",
                title=f"{row.sku.vendor_stock_id}: usable supply or commitment data incomplete",
                effect="Projected ending supply and shortage/surplus cannot be calculated.",
                required_action="Complete current commitments, usable supply and confirmed on-time inbound.",
            )
        else:
            ending = row.usable_supply + row.confirmed_on_time_inbound - row.current_commitments
            row.projected_ending_supply = ending
            row.projected_gap = ending - (row.approved_buffer or Decimal("0"))
            row.decision_status = "INVESTIGATE" if row.projected_gap < 0 else "MONITOR"
            row.save()
    cycle.status = "BLOCKED" if cycle.exceptions.filter(severity="CRITICAL", status="OPEN").exists() else "READY"
    cycle.save(update_fields=["status", "updated_at"])


def refresh_cycle(cycle):
    if cycle.rows.exists():
        run_controls(cycle)
