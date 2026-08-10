from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from .models import ControlCycle, ItemMappingEntry, ReconciliationRow, SKU
from .services import run_controls


class ControlGateTests(TestCase):
    def setUp(self):
        self.cycle = ControlCycle.objects.create(
            name="Test cycle",
            walmart_week="202623",
            cutoff_at=timezone.make_aware(datetime(2026, 7, 24, 20, 40)),
        )
        self.sku = SKU.objects.create(vendor_stock_id="TEST-1", item_name="Test item")
        self.row = ReconciliationRow.objects.create(
            cycle=self.cycle,
            sku=self.sku,
            store_on_hand=Decimal("100"),
        )

    def test_missing_supply_blocks_calculation(self):
        run_controls(self.cycle)
        self.row.refresh_from_db()
        self.cycle.refresh_from_db()
        self.assertEqual(self.cycle.status, "BLOCKED")
        self.assertEqual(self.row.decision_status, "BLOCKED")
        self.assertIsNone(self.row.projected_gap)

    def test_complete_row_calculates_projected_gap(self):
        self.row.current_commitments = Decimal("80")
        self.row.usable_supply = Decimal("100")
        self.row.confirmed_on_time_inbound = Decimal("30")
        self.row.approved_buffer = Decimal("10")
        self.row.save()
        run_controls(self.cycle)
        self.row.refresh_from_db()
        self.assertEqual(self.row.projected_ending_supply, Decimal("50"))
        self.assertEqual(self.row.projected_gap, Decimal("40"))

    def test_main_pages_and_export_render(self):
        self.assertEqual(self.client.get(reverse("controls:dashboard")).status_code, 200)
        video = self.client.get(reverse("controls:demo_video"))
        self.assertEqual(video.status_code, 200)
        self.assertContains(video, "Walmart suppliers face a costly decision")
        self.assertContains(video, "350-unit shortage calculation")
        self.assertContains(video, "Walmart_Control_Beginner_Guide_EN.mp4")
        chinese_video = self.client.get(reverse("controls:demo_video_zh_cn"))
        self.assertEqual(chinese_video.status_code, 200)
        self.assertContains(chinese_video, "沃尔玛供应商面对代价高昂的决策")
        self.assertContains(chinese_video, "350 件预计短缺计算")
        self.assertContains(chinese_video, "Walmart_Control_Beginner_Guide_zh-CN.mp4")
        self.assertEqual(
            self.client.get(reverse("controls:cycle_detail", args=[self.cycle.pk])).status_code,
            200,
        )
        route = self.client.get(reverse("controls:route_control", args=[self.cycle.pk]))
        self.assertEqual(route.status_code, 200)
        self.assertContains(route, "Aggregate evidence by stage")
        self.assertContains(route, "Geography")
        self.assertContains(route, "No operational coordinates are plotted")
        self.assertContains(route, "Unresolved entries")
        detail = self.client.get(reverse("controls:update_decision", args=[self.row.pk]))
        self.assertEqual(detail.status_code, 200)
        self.assertContains(detail, "SINGLE-ITEM ROUTE")
        self.assertContains(detail, "Projected-gap calculation")
        response = self.client.get(reverse("controls:export_excel", args=[self.cycle.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"PK"))

    def test_each_package_has_downloadable_blank_template(self):
        for package_number in range(1, 10):
            response = self.client.get(
                reverse("controls:package_template", args=[package_number])
            )
            self.assertEqual(response.status_code, 200)
            self.assertTrue(response.content.startswith(b"PK"))
            self.assertIn(
                f'filename="P{package_number:02d}_blank_upload_template.xlsx"',
                response["Content-Disposition"],
            )

        factory = self.client.get(reverse("controls:package_template", args=[7]))
        book = load_workbook(BytesIO(factory.content), read_only=True)
        self.assertEqual(
            list(next(book["Data Upload"].iter_rows(values_only=True))),
            [
                "vendor_stock_id",
                "finished_goods_available_quantity",
                "work_in_process_quantity",
                "planned_completion_date",
                "release_date",
            ],
        )
        self.assertIn("Instructions", book.sheetnames)
        book.close()
        mapping = self.client.get(reverse("controls:package_mapping"))
        self.assertEqual(mapping.status_code, 200)
        self.assertContains(mapping, "actual_order_quantity_last_4_weeks")
        self.assertContains(mapping, "otif_exception_reason")
        self.assertContains(mapping, "system_order_start_week")

    def test_expanded_decision_support_calculations_render(self):
        self.row.actual_orders_l4 = Decimal("400")
        self.row.next_week_supply_plan = Decimal("60")
        self.row.otif_on_time_percent = Decimal("0.44")
        self.row.otif_in_full_percent = Decimal("0.569")
        self.row.ecomm_units = Decimal("35")
        self.row.ecomm_on_hand_inventory = Decimal("1831")
        self.row.ecomm_weeks_of_supply = Decimal("52.314")
        self.row.traited_store_count = 1206
        self.row.prior_traited_store_count = 996
        self.row.incremental_store_count = 210
        self.row.initial_fill_units_per_store = Decimal("3")
        self.row.illustrative_initial_fill = Decimal("630")
        self.row.save()
        run_controls(self.cycle)
        self.row.refresh_from_db()
        self.assertEqual(self.row.forecast_variance_units, Decimal("-40"))
        detail = self.client.get(reverse("controls:update_decision", args=[self.row.pk]))
        self.assertContains(detail, "Six forward-decision analyses")
        self.assertContains(detail, "52.3 weeks")
        self.assertContains(detail, "210 incremental stores")

    def test_mapping_review_entries_are_separate_from_decision_rows(self):
        ItemMappingEntry.objects.create(
            cycle=self.cycle,
            internal_sku="DEMO-PROVISIONAL",
            proposed_alias="DEMO-ALIAS",
            gtin="00000000000019",
            status="PROVISIONAL",
            confidence="MEDIUM",
            evidence="Candidate identifier only.",
            required_action="Confirm the Walmart item number.",
        )
        ItemMappingEntry.objects.create(
            cycle=self.cycle,
            internal_sku="DEMO-UNRESOLVED",
            status="UNRESOLVED",
            confidence="LOW",
            evidence="No independent identifier.",
            required_action="Locate catalog evidence.",
        )
        route = self.client.get(reverse("controls:route_control", args=[self.cycle.pk]))
        self.assertContains(route, "DEMO-PROVISIONAL")
        self.assertContains(route, "DEMO-UNRESOLVED")
        self.assertContains(route, "Identifiers outside the decision schedule")
        self.assertNotContains(route, "Synthetic factory")
        self.assertEqual(self.cycle.rows.count(), 1)

    @override_settings(DEMO_READ_ONLY=True, SYNTHETIC_ONLY=True)
    def test_public_demo_hides_real_cycle_and_blocks_changes(self):
        synthetic = ControlCycle.objects.create(
            name="Synthetic cycle",
            walmart_week="202630",
            cutoff_at=timezone.make_aware(datetime(2026, 7, 31, 9, 0)),
            is_synthetic=True,
        )
        dashboard = self.client.get(reverse("controls:dashboard"))
        self.assertContains(dashboard, "Synthetic cycle")
        self.assertNotContains(dashboard, "Test cycle")
        self.assertEqual(
            self.client.get(reverse("controls:cycle_detail", args=[self.cycle.pk])).status_code,
            404,
        )
        self.assertEqual(self.client.get(reverse("controls:cycle_create")).status_code, 403)
        self.assertEqual(
            self.client.post(reverse("controls:run_reconciliation", args=[synthetic.pk])).status_code,
            403,
        )
