from io import BytesIO
from pathlib import Path
from decimal import Decimal

from django.contrib import messages
from django.conf import settings
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import CycleForm, DecisionForm, UploadForm
from .models import ControlCycle, PackageUpload, ReconciliationRow
from .services import process_upload, run_controls
from .template_schemas import PACKAGE_ANALYSIS_USE, PACKAGE_TEMPLATE_SCHEMAS


def _sum_known(rows, field):
    values = [getattr(row, field) for row in rows if getattr(row, field) is not None]
    return sum(values, Decimal("0")) if values else None


def _mapping_summary(sku):
    if sku.walmart_item_number and (sku.gtin or sku.consumer_id or sku.all_links_item_number):
        return "CONFIRMED", "High", "Walmart item plus an independent catalog identifier"
    if sku.walmart_item_number:
        return "CONFIRMED", "Medium", "Walmart item number present; secondary identifier missing"
    if sku.gtin or sku.consumer_id or sku.all_links_item_number:
        return "PROVISIONAL", "Medium", "Candidate identifier present; Walmart item number missing"
    return "UNRESOLVED", "Low", "No Walmart item number or independent catalog identifier"


def dashboard(request):
    cycles = ControlCycle.objects.prefetch_related("uploads", "exceptions").all()
    if settings.SYNTHETIC_ONLY:
        cycles = cycles.filter(is_synthetic=True)
    return render(request, "controls/dashboard.html", {"cycles": cycles})


def demo_video(request, language="en"):
    return render(request, "controls/demo_video.html", {"page_language": language})


def cycle_create(request):
    if settings.DEMO_READ_ONLY:
        return HttpResponseForbidden("This public synthetic demonstration is read-only.")
    form = CycleForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        cycle = form.save(commit=False)
        cycle.prepared_by = "Jessica"
        cycle.save()
        messages.success(request, "Control cycle created.")
        return redirect("controls:cycle_detail", pk=cycle.pk)
    return render(request, "controls/cycle_form.html", {"form": form})


def cycle_detail(request, pk):
    queryset = ControlCycle.objects.all()
    if settings.SYNTHETIC_ONLY:
        queryset = queryset.filter(is_synthetic=True)
    cycle = get_object_or_404(queryset, pk=pk)
    packages = []
    latest = {}
    for upload in cycle.uploads.all():
        latest.setdefault(upload.package_number, upload)
    for number, label in PackageUpload.PACKAGE_CHOICES:
        packages.append({"number": number, "label": label, "upload": latest.get(number)})
    return render(request, "controls/cycle_detail.html", {
        "cycle": cycle,
        "packages": packages,
        "rows": cycle.rows.select_related("sku"),
        "exceptions": cycle.exceptions.select_related("sku"),
    })


def route_control(request, pk):
    queryset = ControlCycle.objects.all()
    if settings.SYNTHETIC_ONLY:
        queryset = queryset.filter(is_synthetic=True)
    cycle = get_object_or_404(queryset, pk=pk)
    reconciliation_rows = list(cycle.rows.select_related("sku"))

    stages = [
        {"key": "factory", "label": "Factory", "field": "work_in_process_quantity", "date": "next_factory_completion", "source": "P07"},
        {"key": "rts", "label": "Ready to ship", "field": "factory_available_inventory", "date": "next_factory_release", "source": "P07"},
        {"key": "staged", "label": "Staged", "field": None, "date": None, "source": "Not modeled"},
        {"key": "waterborne", "label": "Waterborne / inbound", "field": "confirmed_on_time_inbound", "date": "next_eta", "source": "P08"},
        {"key": "port", "label": "Port", "field": None, "date": "next_customs_clearance", "source": "P08 milestone"},
        {"key": "rjw", "label": "3PL warehouse (RJW)", "field": "rjw_physical_inventory", "date": None, "source": "P08"},
        {"key": "dc", "label": "Walmart FC/DC", "field": "ecomm_on_hand_inventory", "date": None, "source": "P03"},
        {"key": "stores", "label": "Walmart stores", "field": "store_on_hand", "date": None, "source": "P01"},
        {"key": "sales", "label": "Sales", "field": "store_pos_units", "date": None, "source": "P01 + P02"},
    ]
    for stage in stages:
        if stage["key"] == "sales":
            store_sales = _sum_known(reconciliation_rows, "store_pos_units")
            ecomm_sales = _sum_known(reconciliation_rows, "ecomm_units")
            known = [value for value in (store_sales, ecomm_sales) if value is not None]
            stage["units"] = sum(known, Decimal("0")) if known else None
            stage["known_count"] = sum(
                1 for row in reconciliation_rows
                if row.store_pos_units is not None or row.ecomm_units is not None
            )
        elif stage["field"]:
            stage["units"] = _sum_known(reconciliation_rows, stage["field"])
            stage["known_count"] = sum(
                1 for row in reconciliation_rows if getattr(row, stage["field"]) is not None
            )
        else:
            stage["units"] = None
            stage["known_count"] = sum(
                1 for row in reconciliation_rows
                if stage["date"] and getattr(row, stage["date"]) is not None
            )
        stage["missing_count"] = len(reconciliation_rows) - stage["known_count"]

    route_rows = []
    mapping_counts = {"CONFIRMED": 0, "PROVISIONAL": 0, "UNRESOLVED": 0}
    for row in reconciliation_rows:
        mapping_status, confidence, mapping_evidence = _mapping_summary(row.sku)
        mapping_counts[mapping_status] += 1
        physical_fields = {
            "factory": row.work_in_process_quantity,
            "rts": row.factory_available_inventory,
            "staged": None,
            "waterborne": row.confirmed_on_time_inbound,
            "port": row.next_customs_clearance,
            "rjw": row.rjw_physical_inventory,
            "dc": row.ecomm_on_hand_inventory,
            "stores": row.store_on_hand,
            "sales": row.store_pos_units if row.store_pos_units is not None else row.ecomm_units,
        }
        known_stages = " ".join(key for key, value in physical_fields.items() if value is not None)
        warnings = []
        if row.confirmed_on_time_inbound and row.rjw_physical_inventory is None:
            warnings.append("Inbound is recorded but the 3PL warehouse physical inventory is not confirmed.")
        if row.rjw_physical_inventory is not None and row.ecomm_on_hand_inventory is None and row.store_on_hand is None:
            warnings.append("3PL warehouse inventory is present but no Walmart inventory position is loaded.")
        if row.next_eta and not row.next_customs_clearance:
            warnings.append("ETA is present but the customs/port milestone is missing.")
        if mapping_status != "CONFIRMED":
            warnings.append("Item identifier mapping requires review before automated reconciliation.")
        route_rows.append({
            "row": row,
            "mapping_status": mapping_status,
            "confidence": confidence,
            "mapping_evidence": mapping_evidence,
            "known_stages": known_stages,
            "warnings": warnings,
        })

    mapping_entries = list(cycle.mapping_entries.all())
    for entry in mapping_entries:
        mapping_counts[entry.status] += 1

    commitments = [
        {"label": "Production", "units": _sum_known(reconciliation_rows, "work_in_process_quantity"), "source": "P07"},
        {"label": "Confirmed inbound", "units": _sum_known(reconciliation_rows, "confirmed_on_time_inbound"), "source": "P08"},
        {"label": "Walmart commitments", "units": _sum_known(reconciliation_rows, "current_commitments"), "source": "P06/P08"},
        {"label": "Order forecast", "units": _sum_known(reconciliation_rows, "order_forecast_total"), "source": "P05"},
        {"label": "Store demand", "units": _sum_known(reconciliation_rows, "forecast_demand_4w"), "source": "P04 · 4 weeks"},
    ]
    geo_locations = []
    geo_segments = []
    if cycle.is_synthetic:
        geo_locations = [
            {"id": "factory", "name": "Synthetic factory", "kind": "Factory", "stage": "factory", "coordinates": [113.2644, 23.1291], "status": "confirmed", "note": "Illustrative Guangdong location—not an operational address."},
            {"id": "origin-port", "name": "Synthetic origin port", "kind": "Port", "stage": "port", "coordinates": [114.2700, 22.5800], "status": "confirmed", "note": "Illustrative export milestone."},
            {"id": "destination-port", "name": "Synthetic destination port", "kind": "Port", "stage": "port", "coordinates": [-118.1900, 33.7500], "status": "confirmed", "note": "Illustrative import milestone."},
            {"id": "rjw", "name": "Synthetic 3PL warehouse (RJW)", "kind": "3PL warehouse", "stage": "rjw", "coordinates": [-90.2000, 38.6300], "status": "provisional", "note": "Made-up demonstration coordinate."},
            {"id": "walmart-dc", "name": "Synthetic Walmart DC", "kind": "Walmart DC", "stage": "dc", "coordinates": [-93.1000, 35.3000], "status": "provisional", "note": "Made-up demonstration coordinate."},
        ]
        geo_segments = [
            {"name": "Factory to origin port", "status": "confirmed", "coordinates": [[113.2644, 23.1291], [114.2700, 22.5800]]},
            {"name": "Pacific shipment", "status": "confirmed", "coordinates": [[114.2700, 22.5800], [179.0, 30.0]]},
            {"name": "Pacific shipment", "status": "confirmed", "coordinates": [[-179.0, 30.0], [-118.1900, 33.7500]]},
            {"name": "Port to synthetic 3PL warehouse", "status": "provisional", "coordinates": [[-118.1900, 33.7500], [-90.2000, 38.6300]]},
            {"name": "Synthetic 3PL warehouse to Walmart DC", "status": "provisional", "coordinates": [[-90.2000, 38.6300], [-93.1000, 35.3000]]},
        ]
    unmapped_locations = [
        "Staged inventory location",
        "Individual Walmart stores",
    ] if cycle.is_synthetic else [
        "Factory and ready-to-ship origin",
        "Origin and destination ports",
        "3PL warehouse facilities",
        "Walmart DC and FC destinations",
        "Individual Walmart stores",
    ]
    return render(request, "controls/route_control.html", {
        "cycle": cycle,
        "stages": stages,
        "commitments": commitments,
        "route_rows": route_rows,
        "mapping_entries": mapping_entries,
        "mapping_counts": mapping_counts,
        "mapping_total": len(reconciliation_rows) + len(mapping_entries),
        "geo_locations": geo_locations,
        "geo_segments": geo_segments,
        "unmapped_locations": unmapped_locations,
    })


def package_upload(request, pk):
    if settings.DEMO_READ_ONLY:
        return HttpResponseForbidden("This public synthetic demonstration is read-only.")
    cycle = get_object_or_404(ControlCycle, pk=pk)
    initial = {"package_number": request.GET.get("package")}
    form = UploadForm(request.POST or None, request.FILES or None, initial=initial)
    if request.method == "POST" and form.is_valid():
        upload = form.save(commit=False)
        upload.cycle = cycle
        upload.original_name = request.FILES["source_file"].name
        upload.save()
        process_upload(upload)
        if upload.status == "ERROR":
            messages.error(request, upload.validation_message)
        else:
            messages.success(request, f"Package {upload.package_number} uploaded: {upload.validation_message}")
        return redirect("controls:cycle_detail", pk=cycle.pk)
    return render(request, "controls/upload_form.html", {"cycle": cycle, "form": form})


def package_template(request, package_number):
    labels = dict(PackageUpload.PACKAGE_CHOICES)
    schema = PACKAGE_TEMPLATE_SCHEMAS.get(package_number)
    if not schema or package_number not in labels:
        return HttpResponse(status=404)

    book = Workbook()
    data = book.active
    data.title = "Data Upload"
    headers = [column for column, _required, _guidance in schema]
    data.append(headers)
    for cell in data[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)
    data.freeze_panes = "A2"
    data.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for index, header in enumerate(headers, 1):
        data.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 3, 18), 42)

    instructions = book.create_sheet("Instructions")
    instructions.append([f"P{package_number:02d}", labels[package_number]])
    instructions.append(["Purpose", "Enter one source record per row on the Data Upload sheet. Do not rename the columns."])
    instructions.append(["Dates", "Use YYYY-MM-DD. Walmart weeks use YYYYYY, for example 202630."])
    instructions.append(["Required fields", "Every required column must remain present; values should be completed for each applicable row."])
    instructions.append([])
    instructions.append(["Column", "Required?", "Meaning / format"])
    for column, required, guidance in schema:
        instructions.append([column, "Yes" if required else "No", guidance])
    for cell in instructions[6]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    instructions.column_dimensions["A"].width = 42
    instructions.column_dimensions["B"].width = 14
    instructions.column_dimensions["C"].width = 78
    instructions.freeze_panes = "A7"

    output = BytesIO()
    book.save(output)
    filename = f"P{package_number:02d}_blank_upload_template.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def package_mapping(request):
    labels = dict(PackageUpload.PACKAGE_CHOICES)
    packages = [
        {
            "number": number,
            "label": labels[number],
            "analysis_use": PACKAGE_ANALYSIS_USE[number],
            "fields": PACKAGE_TEMPLATE_SCHEMAS[number],
        }
        for number in labels
    ]
    return render(request, "controls/package_mapping.html", {"packages": packages})


@require_POST
def run_reconciliation(request, pk):
    if settings.DEMO_READ_ONLY:
        return HttpResponseForbidden("This public synthetic demonstration is read-only.")
    cycle = get_object_or_404(ControlCycle, pk=pk)
    run_controls(cycle)
    messages.success(request, "Reconciliation and exception controls completed.")
    return redirect("controls:cycle_detail", pk=cycle.pk)


def update_decision(request, pk):
    row = get_object_or_404(ReconciliationRow, pk=pk)
    if settings.SYNTHETIC_ONLY and not row.cycle.is_synthetic:
        return HttpResponseForbidden("Only the synthetic demonstration is available.")
    if request.method == "POST" and settings.DEMO_READ_ONLY:
        return HttpResponseForbidden("This public synthetic demonstration is read-only.")
    form = DecisionForm(request.POST or None, instance=row)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"Decision note updated for {row.sku.vendor_stock_id}.")
        return redirect("controls:cycle_detail", pk=row.cycle_id)
    return render(request, "controls/decision_form.html", {"row": row, "form": form})


def export_excel(request, pk):
    queryset = ControlCycle.objects.all()
    if settings.SYNTHETIC_ONLY:
        queryset = queryset.filter(is_synthetic=True)
    cycle = get_object_or_404(queryset, pk=pk)
    book = Workbook()
    sheet = book.active
    sheet.title = "Reconciliation"
    headers = [
        "SKU", "Item", "Store POS Units", "Store POS Sales", "Store On Hand",
        "Store On Order", "Replen In-stock", "eComm Units", "eComm Sales",
        "4W Store Demand", "13W Store Demand", "Order Forecast", "First Arrival",
        "Last Arrival", "Current Commitments", "Usable Supply", "On-time Inbound",
        "Approved Buffer", "Projected Ending", "Projected Gap", "Decision Status",
        "Recommendation", "Decision Note",
    ]
    header_row = 1
    if cycle.is_synthetic:
        sheet.append(["SYNTHETIC DEMONSTRATION DATA — NOT FOR OPERATIONAL USE"])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        sheet["A1"].fill = PatternFill("solid", fgColor="FFF0C9")
        sheet["A1"].font = Font(color="8A5A00", bold=True, size=14)
        sheet.append([])
        header_row = 3
    sheet.append(headers)
    for cell in sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)
    for row in cycle.rows.select_related("sku"):
        sheet.append([
            row.sku.vendor_stock_id, row.sku.item_name, row.store_pos_units, row.store_pos_sales,
            row.store_on_hand, row.store_on_order, row.replen_instock, row.ecomm_units,
            row.ecomm_sales, row.forecast_demand_4w, row.forecast_demand_13w,
            row.order_forecast_total, row.first_forecast_arrival, row.last_forecast_arrival,
            row.current_commitments, row.usable_supply, row.confirmed_on_time_inbound,
            row.approved_buffer, row.projected_ending_supply, row.projected_gap,
            row.get_decision_status_display(), row.recommendation, row.decision_note,
        ])
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 28, 15, 16, 15, 15, 14, 14, 14, 16, 16, 16, 15, 15, 18, 15, 16, 15, 16, 15, 16, 30, 30]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    exceptions = book.create_sheet("Exceptions")
    exceptions.append(["Severity", "SKU", "Code", "Exception", "Effect", "Required Action", "Status", "Note"])
    for cell in exceptions[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    for item in cycle.exceptions.select_related("sku"):
        exceptions.append([
            item.severity, item.sku.vendor_stock_id if item.sku else "", item.code,
            item.title, item.effect, item.required_action, item.status, item.note,
        ])
    decision = book.create_sheet("Decision Support")
    decision_headers = [
        "SKU", "L4 Actual Orders", "Next-week Supply Plan", "Variance Units",
        "Variance Percent", "OTIF On-time", "OTIF In-full", "OTIF Exceptions",
        "eComm On Hand", "eComm FC Count", "eComm Weeks of Supply",
        "WIP", "Factory Release", "ETD", "ETA", "Customs Clearance",
        "Modular Set Week", "System-order Start Week", "Prior Traited Stores",
        "Current Traited Stores", "Incremental Stores", "Initial Fill / Store",
        "Illustrative Initial Fill",
    ]
    decision.append(decision_headers)
    for cell in decision[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in cycle.rows.select_related("sku"):
        decision.append([
            row.sku.vendor_stock_id, row.actual_orders_l4, row.next_week_supply_plan,
            row.forecast_variance_units, row.forecast_variance_percent,
            row.otif_on_time_percent, row.otif_in_full_percent, row.otif_exception_count,
            row.ecomm_on_hand_inventory, row.ecomm_fc_count, row.ecomm_weeks_of_supply,
            row.work_in_process_quantity, row.next_factory_release, row.next_etd, row.next_eta,
            row.next_customs_clearance, row.modular_set_week, row.system_order_start_week,
            row.prior_traited_store_count, row.traited_store_count, row.incremental_store_count,
            row.initial_fill_units_per_store, row.illustrative_initial_fill,
        ])
    decision.freeze_panes = "A2"
    decision.auto_filter.ref = decision.dimensions
    for column in range(1, len(decision_headers) + 1):
        decision.column_dimensions[get_column_letter(column)].width = 20
    output = BytesIO()
    book.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="walmart-control-{cycle.walmart_week}.xlsx"'
    return response


def print_report(request, pk):
    queryset = ControlCycle.objects.all()
    if settings.SYNTHETIC_ONLY:
        queryset = queryset.filter(is_synthetic=True)
    cycle = get_object_or_404(queryset, pk=pk)
    return render(request, "controls/print_report.html", {
        "cycle": cycle,
        "rows": cycle.rows.select_related("sku"),
        "exceptions": cycle.exceptions.select_related("sku"),
    })
